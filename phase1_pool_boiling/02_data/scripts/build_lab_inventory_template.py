"""build_lab_inventory_template.py — Generate empty Phase 1.5 lab data inventory workbook.

Creates ``phase1p5_inhouse_augmentation/data/lab_data_inventory.xlsx`` with
seven sheets pre-populated with column headers, data validation drop-downs,
cell comments explaining tricky fields, frozen header rows, and example
rows that graduate students can use as a guide.

Run once to (re)build the template. Do not run after students have started
filling it in — it would overwrite their data.

Usage:
    python 02_data/scripts/build_lab_inventory_template.py
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = (
    ROOT
    / "phase1p5_inhouse_augmentation"
    / "data"
    / "lab_data_inventory.xlsx"
)

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow for required fields
EXAMPLE_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green for example rows
EXAMPLE_FONT = Font(italic=True, color="385723")


def write_headers(ws, headers: list[str], comments: dict[str, str] | None = None) -> None:
    """Apply header row 1 styling + optional cell comments."""
    comments = comments or {}
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        if name in comments:
            cell.comment = Comment(comments[name], "Phase 1.5 template")
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"


def set_column_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ---------------------------------------------------------------------------
# Drop-down lists (used in multiple sheets)
# ---------------------------------------------------------------------------
FLUID_OPTIONS = '"water,R-123,R-134a"'   # FC-77 등은 PINN 미지원
MATERIAL_OPTIONS = '"Cu,Al,SS,Brass,Si,glass,bronze,other"'
STATUS_OPTIONS = '"TBD,in_progress,complete,blocked,N/A"'
TIER_OPTIONS = '"A,B,C,D"'
YN_OPTIONS = '"Yes,No,partial"'
GEOMETRY_OPTIONS = '"flat_plate_horizontal_up,flat_plate_horizontal_down,cylinder_horizontal,wire,other"'


# ===========================================================================
# Sheet 1 — Overview
# ===========================================================================
def make_overview_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Overview", 0)
    headers = [
        "source",
        "doi",
        "status",
        "n_surfaces_target",
        "n_surfaces_collected",
        "n_curves_target",
        "n_curves_collected",
        "n_onb_target",
        "n_onb_collected",
        "raw_data_format",
        "sem_available",
        "data_quality_tier",
        "responsible_student",
        "deadline_M1_M3",
        "notes",
    ]
    comments = {
        "status": "TBD / in_progress / complete / blocked / N/A",
        "raw_data_format": "CSV / Excel / paper_figure_only / mixed",
        "sem_available": "Yes / No / partial (top-view only, etc.)",
        "data_quality_tier": "A (raw CSV + visual ONB) / B (raw CSV + slope) / "
                             "C (figure digitize) / D (uncertain)",
        "deadline_M1_M3": "예: 2026-07-15 (M1 end), 2026-09-30 (M3 end)",
    }
    write_headers(ws, headers, comments)

    set_column_widths(ws, {
        "A": 24, "B": 32, "C": 13, "D": 16, "E": 18, "F": 16, "G": 18,
        "H": 16, "I": 18, "J": 18, "K": 13, "L": 13, "M": 22, "N": 22, "O": 38,
    })

    # Pre-populated 4 source rows (empty fields TBD)
    rows = [
        ["Lee_2023_ICHMT", "10.1016/j.icheatmasstransfer.2023.107072",
         "TBD", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Lee_2024_ICHMT", "10.1016/j.icheatmasstransfer.2024.107270",
         "TBD", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Inhouse_corrosion", "(unpublished)",
         "TBD", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Inhouse_biphilic",  "(unpublished)",
         "TBD", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["Total",             "",
         "—",   "", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 1 and val == "Total":
                cell.font = Font(bold=True)

    # Data validation for status column (C)
    dv_status = DataValidation(type="list", formula1=STATUS_OPTIONS, allow_blank=True)
    dv_status.add("C2:C5")
    ws.add_data_validation(dv_status)

    # Data validation for raw_data_format (J)
    dv_raw = DataValidation(
        type="list",
        formula1='"CSV,Excel,paper_figure_only,mixed,no_data"',
        allow_blank=True,
    )
    dv_raw.add("J2:J5")
    ws.add_data_validation(dv_raw)

    # SEM available (K)
    dv_sem = DataValidation(type="list", formula1=YN_OPTIONS, allow_blank=True)
    dv_sem.add("K2:K5")
    ws.add_data_validation(dv_sem)

    # Data quality tier (L)
    dv_tier = DataValidation(type="list", formula1=TIER_OPTIONS, allow_blank=True)
    dv_tier.add("L2:L5")
    ws.add_data_validation(dv_tier)


# ===========================================================================
# Sheets 2-5 — Per-source surface metadata
# ===========================================================================
def make_source_sheet(wb: Workbook, source_name: str, example_row: list[Any] | None) -> None:
    ws = wb.create_sheet(source_name)
    headers = [
        "surface_id",            # SFC-051, ...
        "surface_label",         # Cu_laser_F100
        "material",              # Cu / Al / SS / Brass / Si / glass / bronze
        "treatment_primary",     # laser_ripple / Cu_sinter+laser / corrosion / biphilic_pattern
        "treatment_parameters",  # laser fluence, etch time, etc. (free text)
        "Ra_um",
        "Ra_measurement_method", # AFM / stylus / SEM_image_analysis / reported_band
        "theta_static_deg",
        "theta_measurement_method",  # sessile_drop / dynamic
        "fluid",
        "pressure_kPa",
        "subcooling_K",
        "heater_geometry",
        "heater_size_mm",
        "n_boiling_curves",
        "n_onb_labels_extractable",
        "sem_top_view_filename",
        "sem_cross_section_filename",
        "N_s_per_cm2",           # cavity density (optional, SEM-derived)
        "r_c_distribution_um",   # SEM 측정 r_c (optional)
        "data_quality_tier",     # A/B/C/D
        "experiment_date",
        "experimenter",
        "thesis_used_in",
        "notes",
    ]
    comments = {
        "surface_id": "SFC-051 부터 시작. Phase 1.5 신규 ID. _index.md 갱신과 일치 필요.",
        "surface_label": "사람이 이해 가능한 짧은 이름. 예: Cu_laser_F100, Cu_sinter120_laser, Cu_corr_long",
        "material": "drop-down: Cu / Al / SS / Brass / Si / glass / bronze / other",
        "treatment_primary": "예: laser_ripple, Cu_sinter+laser, corrosion_oxide, biphilic_SiO2_Teflon",
        "treatment_parameters": "자유 텍스트. 예: 'fluence=20 J/cm², 100 pulses', 'etching 30 min in NaCl 3%'",
        "Ra_um": "단위 반드시 μm. nm 측정값은 ÷1000. 예: 30 nm → 0.030",
        "Ra_measurement_method": "AFM / stylus / SEM_image_analysis / reported_band",
        "theta_static_deg": "단위 °. 미측정 시 빈칸.",
        "theta_measurement_method": "sessile_drop / dynamic / not_measured",
        "fluid": "drop-down: water / R-123 / R-134a (FC-77/HFE-7100은 CoolProp 미지원으로 제외)",
        "pressure_kPa": "예: 101.325 (water at 1 atm), 500 (R-134a 보통 5 bar)",
        "subcooling_K": "T_sat - T_bulk. Saturated test면 0.",
        "heater_geometry": "drop-down: flat_plate_horizontal_up / cylinder_horizontal / wire / other",
        "heater_size_mm": "예: 10×10 mm² 평판이면 '10×10', 원통이면 'D=4, L=20'",
        "n_boiling_curves": "이 surface에서 측정된 boiling curve 수",
        "n_onb_labels_extractable": "ONB 식별 가능한 라벨 수 (보통 1-3 per curve)",
        "sem_top_view_filename": "예: SFC-051_top_5kx.png — 실제 파일은 data/raw/lab/<source>/sem/ 에 보관",
        "sem_cross_section_filename": "단면 SEM이 있다면 파일명",
        "N_s_per_cm2": "공동 밀도 (SEM 정량 분석). 없으면 빈칸",
        "r_c_distribution_um": "활성 공동 반경 분포. 예: 'median=3.2 μm, IQR=[1.5, 6.8]'",
        "data_quality_tier": "A (raw CSV + visual ONB) / B (raw CSV + slope) / C (figure digitize) / D (uncertain)",
        "experiment_date": "YYYY-MM-DD",
        "experimenter": "측정자 이름 (예: 이승환, 김연수)",
        "thesis_used_in": "해당 학위논문 (예: 이승환 박사 2024) 또는 빈칸",
        "notes": "기타 — 공동저자 우려, 재측정 필요, 데이터 staleness 등",
    }
    write_headers(ws, headers, comments)

    set_column_widths(ws, {
        "A": 13, "B": 25, "C": 12, "D": 24, "E": 30, "F": 10, "G": 22,
        "H": 16, "I": 22, "J": 12, "K": 14, "L": 14, "M": 25, "N": 18,
        "O": 17, "P": 22, "Q": 26, "R": 26, "S": 14, "T": 22, "U": 13,
        "V": 16, "W": 16, "X": 25, "Y": 38,
    })

    # Data validations
    dv_material = DataValidation(type="list", formula1=MATERIAL_OPTIONS, allow_blank=True)
    dv_material.add("C2:C200")
    ws.add_data_validation(dv_material)

    dv_fluid = DataValidation(type="list", formula1=FLUID_OPTIONS, allow_blank=True)
    dv_fluid.add("J2:J200")
    ws.add_data_validation(dv_fluid)

    dv_geom = DataValidation(type="list", formula1=GEOMETRY_OPTIONS, allow_blank=True)
    dv_geom.add("M2:M200")
    ws.add_data_validation(dv_geom)

    dv_tier = DataValidation(type="list", formula1=TIER_OPTIONS, allow_blank=True)
    dv_tier.add("U2:U200")
    ws.add_data_validation(dv_tier)

    # Example row (italic green)
    if example_row is not None:
        for c_idx, val in enumerate(example_row, start=1):
            cell = ws.cell(row=2, column=c_idx, value=val)
            cell.fill = EXAMPLE_FILL
            cell.font = EXAMPLE_FONT
        # mark example row in column Y (notes) explicitly
        ws.cell(row=2, column=len(headers),
                value="(예시 행 — 실제 데이터 입력 시 이 행을 지우거나 위에 덮어쓰기)")


# ===========================================================================
# Sheet 6 — Boiling curves (long format, schema = Phase 1 boiling_curves.csv)
# ===========================================================================
def make_boiling_curves_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Boiling_curves")
    headers = [
        "source_paper",      # Lee_2023_ICHMT, etc.
        "figure_ref",        # run01, Fig3a, etc.
        "surface_id",        # SFC-051
        "surface_label",     # Cu_laser_F100
        "fluid",
        "delta_T_wall",      # K
        "delta_T_sub",       # K
        "q_flux",            # W/m²
        "Ra_um",
        "theta_deg",
        "category",
        "ONB_flag",          # TRUE / FALSE
        "r_c_um",            # inferred (only for ONB rows)
        "notes",
    ]
    comments = {
        "source_paper": "Lee_2023_ICHMT / Lee_2024_ICHMT / Inhouse_corrosion / Inhouse_biphilic",
        "figure_ref": "원본 figure 또는 run 식별자. 예: Fig3a, trial_05",
        "fluid": "water / R-123 / R-134a (drop-down)",
        "delta_T_wall": "단위 K. T_wall - T_sat",
        "delta_T_sub": "단위 K. T_sat - T_bulk. Saturated이면 0",
        "q_flux": "단위 W/m². kW/m²면 ×1000 변환 후 입력",
        "Ra_um": "Surface별 metadata에서 자동 채워질 수 있음",
        "theta_deg": "동일",
        "category": "betz / bourdon12 / ... / unist_laser / unist_corrosion / unist_biphilic",
        "ONB_flag": "TRUE = 이 점이 ONB / FALSE = 일반 boiling curve 점",
        "r_c_um": "Hsu inverse 자동 계산 값 (ONB rows only). 빈칸 OK",
        "notes": "ONB_manual / ONB_auto / 측정 비고",
    }
    write_headers(ws, headers, comments)

    set_column_widths(ws, {
        "A": 22, "B": 16, "C": 13, "D": 25, "E": 12,
        "F": 14, "G": 14, "H": 14, "I": 10, "J": 12,
        "K": 18, "L": 12, "M": 12, "N": 38,
    })

    # Data validations
    dv_fluid = DataValidation(type="list", formula1=FLUID_OPTIONS, allow_blank=True)
    dv_fluid.add("E2:E5000")
    ws.add_data_validation(dv_fluid)

    dv_source = DataValidation(
        type="list",
        formula1='"Lee_2023_ICHMT,Lee_2024_ICHMT,Inhouse_corrosion,Inhouse_biphilic"',
        allow_blank=True,
    )
    dv_source.add("A2:A5000")
    ws.add_data_validation(dv_source)

    dv_onb = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv_onb.add("L2:L5000")
    ws.add_data_validation(dv_onb)

    # Single example row (green italic)
    example = [
        "Lee_2023_ICHMT", "Fig3a_F100", "SFC-051", "Cu_laser_F100",
        "water", 5.2, 0.0, 50000, 0.5, 85.0, "unist_laser", "TRUE", "",
        "예시 — 실제 입력 시 본 행 삭제 또는 덮어쓰기",
    ]
    for c_idx, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=c_idx, value=val)
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT


# ===========================================================================
# Sheet 7 — ONB labels (subset, schema = Phase 1 onb_dataset.csv)
# ===========================================================================
def make_onb_labels_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("ONB_labels")
    headers = [
        "source_paper", "figure_ref", "surface_id", "surface_label", "fluid",
        "delta_T_wall", "delta_T_sub", "q_flux", "Ra_um", "theta_deg",
        "category", "notes",
    ]
    comments = {
        "delta_T_wall": "ONB 시점의 wall 과열도",
        "q_flux": "ONB 시점의 q'' (W/m²)",
        "notes": "ONB_manual / ONB_slope_auto / 시각화 확인 / tier 정보",
    }
    write_headers(ws, headers, comments)

    set_column_widths(ws, {
        "A": 22, "B": 16, "C": 13, "D": 25, "E": 12,
        "F": 14, "G": 14, "H": 14, "I": 10, "J": 12,
        "K": 18, "L": 38,
    })

    # Same drop-downs
    dv_fluid = DataValidation(type="list", formula1=FLUID_OPTIONS, allow_blank=True)
    dv_fluid.add("E2:E1000")
    ws.add_data_validation(dv_fluid)
    dv_source = DataValidation(
        type="list",
        formula1='"Lee_2023_ICHMT,Lee_2024_ICHMT,Inhouse_corrosion,Inhouse_biphilic"',
        allow_blank=True,
    )
    dv_source.add("A2:A1000")
    ws.add_data_validation(dv_source)


# ===========================================================================
# Build workbook
# ===========================================================================
def build() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    make_overview_sheet(wb)

    # Per-source sheets with one example row each (forward-looking, green-italic)
    make_source_sheet(
        wb, "Lee_2023_ICHMT",
        example_row=[
            "SFC-051", "Cu_laser_F20", "Cu", "laser_ripple",
            "fluence=20 J/cm², 100 pulses, λ=1030nm", 0.45, "AFM",
            85.0, "sessile_drop",
            "water", 101.325, 0.0,
            "flat_plate_horizontal_up", "10×10 mm²",
            6, 2,
            "SFC-051_top_5kx.png", "(없음)",
            "", "median=2.8 μm, IQR=[1.2, 5.0]",
            "B", "2023-XX-XX", "Seunghwan Lee",
            "(Lee 박사 학위논문 — 본 데이터 published)",
            "예시 행 — 실제 데이터 채울 때 본 행을 위에 덮어쓰거나 지움",
        ],
    )
    make_source_sheet(
        wb, "Lee_2024_ICHMT",
        example_row=[
            "SFC-058", "Cu_sinter120_laser_F20", "Cu", "Cu_sinter+laser",
            "sinter t=120μm + laser fluence 20 J/cm²", 0.8, "AFM",
            "", "(not measured)",
            "water", 101.325, 0.0,
            "flat_plate_horizontal_up", "20×20 mm²",
            5, 1,
            "SFC-058_top_5kx.png", "(없음)",
            "", "(SEM 측정 미실시)",
            "B", "2024-XX-XX", "Seunghwan Lee",
            "(Lee 박사 학위논문)",
            "예시 행",
        ],
    )
    make_source_sheet(
        wb, "Inhouse_corrosion",
        example_row=[
            "SFC-063", "Cu_corr_60min", "Cu", "corrosion_oxide",
            "etching 60 min in NaCl 3%", 0.05, "stylus",
            22.0, "sessile_drop",
            "water", 101.325, 0.0,
            "flat_plate_horizontal_up", "10×10 mm²",
            4, 2,
            "(미측정)", "(미측정)",
            "", "",
            "C", "(연도 미상)", "(측정자 확인 필요)",
            "(미공개 — 본 paper 첫 publication 후보)",
            "예시 행",
        ],
    )
    make_source_sheet(
        wb, "Inhouse_biphilic",
        example_row=[
            "SFC-067", "biphilic_SiO2_Teflon_50um", "Si", "biphilic_pattern",
            "SiO2 background + Teflon dots, 50 μm pitch", 0.002, "AFM",
            "mixed_pattern", "sessile_drop_zones",
            "water", 101.325, 0.0,
            "flat_plate_horizontal_up", "10×10 mm²",
            3, 1,
            "SFC-067_top_10kx.png", "(없음)",
            "", "",
            "B", "(연도 미상)", "(측정자 확인 필요)",
            "(미공개 — 본 paper 첫 publication 후보)",
            "예시 행",
        ],
    )

    make_boiling_curves_sheet(wb)
    make_onb_labels_sheet(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"✓ Inventory template written: {OUT_PATH}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    build()
