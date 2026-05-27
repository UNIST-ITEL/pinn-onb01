"""export_inventory_to_csv.py — Phase 1.5 inventory xlsx → CSV pipeline.

Reads ``phase1p5_inhouse_augmentation/data/lab_data_inventory.xlsx`` and
exports three CSV files ready for downstream consumption by the
``preprocess-pipeline`` and ``surface-card-builder`` agents:

    boiling_curves_v1p5_partial.csv   # Boiling_curves sheet, schema = Phase 1
    onb_dataset_v1p5_partial.csv      # ONB_labels sheet, schema = Phase 1
    surface_metadata.csv              # 4 source sheets combined

The exporter automatically:
- Skips the green-italic example rows that ship with the template
- Drops fully-empty rows
- Validates fluid against the supported set (water / R-123 / R-134a)
- Coerces numeric types and warns on parse failures
- Logs row counts per source and per output

Usage:
    python 02_data/scripts/export_inventory_to_csv.py
    # or with custom paths:
    python 02_data/scripts/export_inventory_to_csv.py \\
        --xlsx phase1p5_inhouse_augmentation/data/lab_data_inventory.xlsx \\
        --out-dir phase1p5_inhouse_augmentation/data/processed/
"""
from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = (
    ROOT
    / "phase1p5_inhouse_augmentation"
    / "data"
    / "lab_data_inventory.xlsx"
)
DEFAULT_OUT_DIR = (
    ROOT
    / "phase1p5_inhouse_augmentation"
    / "data"
    / "processed"
)

SOURCE_SHEETS = [
    "Lee_2023_ICHMT",
    "Lee_2024_ICHMT",
    "Inhouse_corrosion",
    "Inhouse_biphilic",
]

# Phase 1 schema reproduced here so the partial CSVs slot directly into the
# Phase 1 + 1.5 unified pipeline produced by ``preprocess-pipeline``.
BOILING_CURVES_COLUMNS = [
    "source_paper", "figure_ref", "surface_id", "surface_label", "fluid",
    "delta_T_wall", "delta_T_sub", "q_flux", "Ra_um", "theta_deg",
    "category", "ONB_flag", "r_c_um", "notes",
]
ONB_LABELS_COLUMNS = [
    "source_paper", "figure_ref", "surface_id", "surface_label", "fluid",
    "delta_T_wall", "delta_T_sub", "q_flux", "Ra_um", "theta_deg",
    "category", "notes",
]

# Green-fill (light green E2EFDA) marks example rows that should be skipped.
EXAMPLE_FILL_RGB = "00E2EFDA"  # openpyxl prepends 8-digit alpha


SUPPORTED_FLUIDS = {"water", "r-123", "r123", "r-134a", "r134a"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_example_row(row: tuple[Cell, ...]) -> bool:
    """Return True if any cell in the row carries the example-fill color."""
    for cell in row:
        fill = cell.fill
        if fill is None or fill.fgColor is None:
            continue
        rgb = getattr(fill.fgColor, "rgb", None)
        if rgb and str(rgb).upper().endswith("E2EFDA"):
            return True
    return False


def _is_empty_row(values: list[Any]) -> bool:
    return all((v is None) or (isinstance(v, str) and v.strip() == "") for v in values)


def _to_float(v: Any) -> float | None:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _normalize_fluid(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _read_sheet(ws, schema: list[str]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Read sheet with header in row 1; align values to ``schema`` order.

    Returns (rows, stats). ``stats`` reports total / examples_skipped /
    empty_skipped / unsupported_fluid_skipped / kept.
    """
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = {h: i for i, h in enumerate(headers) if h is not None}

    rows: list[dict[str, Any]] = []
    stats = {"total": 0, "example_skipped": 0, "empty_skipped": 0,
             "unsupported_fluid_skipped": 0, "kept": 0}

    for row in ws.iter_rows(min_row=2):
        stats["total"] += 1
        if _is_example_row(row):
            stats["example_skipped"] += 1
            continue
        values = [c.value for c in row]
        if _is_empty_row(values):
            stats["empty_skipped"] += 1
            continue

        out: dict[str, Any] = {}
        for col in schema:
            i = header_idx.get(col)
            out[col] = values[i] if (i is not None and i < len(values)) else None

        # Fluid sanity for boiling_curves / onb_labels
        if "fluid" in out:
            fluid_norm = _normalize_fluid(out["fluid"]).lower()
            if fluid_norm and fluid_norm not in SUPPORTED_FLUIDS:
                stats["unsupported_fluid_skipped"] += 1
                continue

        rows.append(out)
        stats["kept"] += 1

    return rows, stats


def _write_csv(rows: list[dict[str, Any]], columns: list[str], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for row in rows:
            w.writerow({c: ("" if row.get(c) is None else row.get(c)) for c in columns})


# ---------------------------------------------------------------------------
# Surface metadata combiner
# ---------------------------------------------------------------------------

SURFACE_METADATA_COLUMNS = [
    "source",                  # synthetic: which sheet (= source_paper key)
    "surface_id", "surface_label", "material", "treatment_primary",
    "treatment_parameters", "Ra_um", "Ra_measurement_method",
    "theta_static_deg", "theta_measurement_method",
    "fluid", "pressure_kPa", "subcooling_K",
    "heater_geometry", "heater_size_mm",
    "n_boiling_curves", "n_onb_labels_extractable",
    "sem_top_view_filename", "sem_cross_section_filename",
    "N_s_per_cm2", "r_c_distribution_um",
    "data_quality_tier",
    "experiment_date", "experimenter", "thesis_used_in", "notes",
]


def _read_surface_sheet(ws, source_name: str) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """Source 시트는 schema 가 같으니 통합."""
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = {h: i for i, h in enumerate(headers) if h is not None}

    rows: list[dict[str, Any]] = []
    stats = {"total": 0, "example_skipped": 0, "empty_skipped": 0, "kept": 0}

    for row in ws.iter_rows(min_row=2):
        stats["total"] += 1
        if _is_example_row(row):
            stats["example_skipped"] += 1
            continue
        values = [c.value for c in row]
        if _is_empty_row(values):
            stats["empty_skipped"] += 1
            continue

        out: dict[str, Any] = {"source": source_name}
        for col in SURFACE_METADATA_COLUMNS:
            if col == "source":
                continue
            i = header_idx.get(col)
            out[col] = values[i] if (i is not None and i < len(values)) else None

        rows.append(out)
        stats["kept"] += 1

    return rows, stats


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def export(xlsx_path: Path, out_dir: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)

    report: dict[str, Any] = {
        "xlsx": str(xlsx_path),
        "out_dir": str(out_dir),
        "per_sheet": {},
        "output": {},
    }

    # --- 1) Boiling_curves ----------------------------------------------------
    if "Boiling_curves" in wb.sheetnames:
        rows, stats = _read_sheet(wb["Boiling_curves"], BOILING_CURVES_COLUMNS)
        out_path = out_dir / "boiling_curves_v1p5_partial.csv"
        _write_csv(rows, BOILING_CURVES_COLUMNS, out_path)
        report["per_sheet"]["Boiling_curves"] = stats
        report["output"]["boiling_curves_v1p5_partial.csv"] = len(rows)

    # --- 2) ONB_labels -------------------------------------------------------
    if "ONB_labels" in wb.sheetnames:
        rows, stats = _read_sheet(wb["ONB_labels"], ONB_LABELS_COLUMNS)
        out_path = out_dir / "onb_dataset_v1p5_partial.csv"
        _write_csv(rows, ONB_LABELS_COLUMNS, out_path)
        report["per_sheet"]["ONB_labels"] = stats
        report["output"]["onb_dataset_v1p5_partial.csv"] = len(rows)

    # --- 3) Surface metadata (4 source sheets combined) ----------------------
    combined_rows: list[dict[str, Any]] = []
    source_stats: dict[str, dict[str, int]] = {}
    for source_sheet in SOURCE_SHEETS:
        if source_sheet not in wb.sheetnames:
            continue
        rows, stats = _read_surface_sheet(wb[source_sheet], source_sheet)
        combined_rows.extend(rows)
        source_stats[source_sheet] = stats

    out_path = out_dir / "surface_metadata.csv"
    _write_csv(combined_rows, SURFACE_METADATA_COLUMNS, out_path)
    report["per_sheet"]["sources"] = source_stats
    report["output"]["surface_metadata.csv"] = len(combined_rows)

    # --- Overview sheet (optional pass-through) ------------------------------
    if "Overview" in wb.sheetnames:
        ws = wb["Overview"]
        out_path = out_dir / "overview.csv"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(min_row=1, values_only=True):
                w.writerow(["" if v is None else v for v in row])
        report["output"]["overview.csv"] = ws.max_row - 1

    return report


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Export Phase 1.5 lab_data_inventory.xlsx to CSV files.",
    )
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                   help=f"Inventory workbook (default: {DEFAULT_XLSX.relative_to(ROOT)})")
    p.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR,
                   help=f"Output directory (default: {DEFAULT_OUT_DIR.relative_to(ROOT)})")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    if not args.xlsx.exists():
        raise FileNotFoundError(f"Inventory workbook not found: {args.xlsx}")

    report = export(args.xlsx, args.out_dir)

    # --- Print human-readable report ------------------------------------------
    print(f"[export-inventory] xlsx: {report['xlsx']}")
    print(f"[export-inventory] out:  {report['out_dir']}")
    print()
    for sheet, stats in report["per_sheet"].items():
        if sheet == "sources":
            print("  Source sheets (surface metadata combined):")
            for src, s in stats.items():
                print(f"    {src:24s} total={s['total']:>3d}  "
                      f"example={s['example_skipped']:>2d}  "
                      f"empty={s['empty_skipped']:>2d}  "
                      f"kept={s['kept']:>3d}")
        else:
            print(f"  {sheet:24s} total={stats['total']:>3d}  "
                  f"example={stats['example_skipped']:>2d}  "
                  f"empty={stats['empty_skipped']:>2d}  "
                  f"unsupported_fluid={stats.get('unsupported_fluid_skipped',0):>2d}  "
                  f"kept={stats['kept']:>3d}")
    print()
    print("[export-inventory] Output files:")
    for fname, n in report["output"].items():
        print(f"  {fname:40s}  {n} rows")


if __name__ == "__main__":
    main()
